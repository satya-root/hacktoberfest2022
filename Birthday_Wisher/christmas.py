# --------------------------     Birthday Wisher      ---------------
# -------------------------- Developed by @satya-root ---------------

import openpyxl
from twilio.rest import Client
import time
import datetime

workbook = openpyxl.load_workbook(
    '/home/kali/Desktop/VS Code/Python/Birthday_Wisher/christmas.xlsx')
active = workbook.active

def convert_to_secs():
    current_hr = int(time.strftime("%H"))       # +5 for Heroku
    current_min = int(time.strftime("%M"))      # +30 for Heroku
    current_secs = int(time.strftime("%S"))

    # Following lines are only valid for Heroku
    # while (current_min >= 60):
    #     current_min-=60
    #     current_hr+=1
        
    # while (current_hr == 24):
    #     day = int(date.day) + 1
    #     str_date = str(day) + '_0' + str(date.month)
    #     current_hr = 00

    # -----   Total 86400 secs in 1 day     -----
    # -----       1 hr = 3600secs           -----
    secs = (current_hr * 3600) + (current_min * 60) + (current_secs)
    remaining_secs = 86400 - secs
    print(current_hr)
    print(current_min)
    return remaining_secs


date = datetime.date.today()
str_date = str(date.day) + '_' + str(date.month)
print(str_date)

z = convert_to_secs()
print('remaining secs = ', z)

if (z != 0):
    # ----- Twilio Declarations -----
    account_sid = 'ACa4fe6fb7c47847a1d34d53f7d18a76db'
    auth_token = '0a510259a16b6b8f44e6c6a0fd7ac0e7'
    client = Client(account_sid, auth_token)
    time.sleep(z)

    # ----- Excel Sheet Iteration -----
    for i in range(2, active.max_row+1):
        birthday = str(active.cell(row=i, column=2).value)
        # print(birthday)

        if (str_date == birthday):
            ph = str(int(active.cell(row=i, column=3).value))
            wishes = str(active.cell(row=i, column=4).value)
            # ----- Uncomment next line to see your list of phone numbers and wishes-----
            # print('+91' + ph)
            print(wishes)

            # ----- Send Message -----
            message = client.messages.create(
                from_='+12058929610',
                body=wishes,
                to='+91' + ph
            )
            print(message.sid)

        # rerun = sys.executable    
        # os.execl(rerun, rerun, * sys.argv)
        


# Heroku worker dyno automatically reruns the program when it stops/exits/ends.

# -------------------- End -----------------------
